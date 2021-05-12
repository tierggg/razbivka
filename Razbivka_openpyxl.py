import openpyxl
import xlsxwriter

workbook = openpyxl.load_workbook('test.xlsx')
worksheet = workbook.active

rows = worksheet.max_row
cols = worksheet.max_column

print('строк =', rows, 'столбцов =', cols)

regColumn = 3
firstReg = 2
regNumber = set()

for row in range(firstReg, rows+1):
    if worksheet.cell(row, regColumn).value is not None:  # Проверка пуста ли ячейка
        regNumber.add(worksheet.cell(row, regColumn).value)
for val in regNumber:
    print(val)
print('Найдено', len(regNumber), 'рег. номеров')

for val in regNumber:

    if isinstance(val, float):  # Иногда рег.номера читаются как float
        valname = str(int(val))
    else:
        valname = str(val)

    outbook = xlsxwriter.Workbook('D:\\1\\out\\'+valname+'.xlsx')
    outsheet = outbook.add_worksheet(valname)









    outbook.close()